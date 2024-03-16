import openpyxl as xl

#load work book
wb=xl.load_workbook("D:\\capgemini\\Book1.xlsx")
sh=wb['Sheet1']
print(sh.title)
rows=sh.max_row
columns=sh.max_column
print("total number of rows are "+str(rows))
print("total number of columns are "+str(columns))
for r in range(1,rows+1):
    for c in range(1,columns+1):
        v=sh.cell(r,c)
        print(v.value)

